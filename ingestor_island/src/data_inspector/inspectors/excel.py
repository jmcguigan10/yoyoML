from __future__ import annotations

import io
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook  # type: ignore

from ..core.context import InspectionContext
from ..core.report import DetectionResult
from ..core.source import DataSource, PathSource
from .base import BaseInspector

try:
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore


class ExcelInspector(BaseInspector):
    def _inspect(
        self, source: DataSource, detection: DetectionResult, ctx: InspectionContext
    ) -> Tuple[Dict[str, Any], List[str]]:
        warnings: List[str] = []

        # Load workbook (read-only to be kinder to memory)
        if isinstance(source, PathSource):
            wb = load_workbook(filename=str(source.path), read_only=True, data_only=True)
        else:
            with source.open_binary() as f:
                data = f.read(ctx.max_bytes_to_load)
            wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)

        sheet_summaries: List[Dict[str, Any]] = []
        for name in wb.sheetnames[: min(len(wb.sheetnames), 20)]:
            ws = wb[name]
            max_row = ws.max_row
            max_col = ws.max_column

            # Sample values
            rows = []
            for r in range(1, min(max_row, ctx.max_rows) + 1):
                row_vals = []
                for c in range(1, min(max_col, ctx.max_cols) + 1):
                    row_vals.append(ws.cell(row=r, column=c).value)
                rows.append(row_vals)

            sheet_summary: Dict[str, Any] = {
                "name": name,
                "used_range": {"rows": int(max_row), "cols": int(max_col)},
                "sample_rows": rows,
            }

            # If pandas is available, try to parse this sheet for nicer column names and sample records
            if pd is not None:
                try:
                    if isinstance(source, PathSource):
                        df = pd.read_excel(source.path, sheet_name=name, nrows=ctx.max_rows, engine="openpyxl")
                    else:
                        with source.open_binary() as f:
                            data = f.read(ctx.max_bytes_to_load)
                        df = pd.read_excel(io.BytesIO(data), sheet_name=name, nrows=ctx.max_rows, engine="openpyxl")
                    sheet_summary["columns"] = [str(c) for c in list(df.columns)[: ctx.max_cols]]
                    sheet_summary["sample_records"] = df.head(ctx.max_rows).to_dict(orient="records")
                except Exception as e:
                    warnings.append(f"pandas couldn't parse sheet '{name}': {e.__class__.__name__}: {e}")

            sheet_summaries.append(sheet_summary)

        summary: Dict[str, Any] = {
            "format": "xlsx",
            "sheets": sheet_summaries,
            "sheet_count": len(wb.sheetnames),
        }

        if len(wb.sheetnames) > 20:
            warnings.append("Showing only first 20 sheets.")

        return summary, warnings
